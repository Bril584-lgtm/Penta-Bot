"""Auto-sync Rocket Rivals Season 8 schedules from the official Google Sheets.

Every 6 hours (and on demand via /syncschedule) the cog downloads the three
division sheets as xlsx, re-parses the matchup grids, and hot-reloads
rr_schedule if anything changed. Matchups in the sheets are banner images,
so teams are identified by image content hash (rr_teams.json). Unknown
banners (e.g. a brand-new team) are logged and their matches skipped until
the mapping is updated.
"""

import hashlib
import io
import json
import os
import re
import zipfile
from collections import defaultdict
from datetime import datetime

import aiohttp
import discord
from discord import app_commands
from discord.ext import commands, tasks

import rr_schedule
from statepath import state_file

SHEET_IDS = {
    "Challengers": "11Ma9369xpbWUzivr8CYwahte6IVxGcBfAOCCgxj6i9U",
    "Legends": "1VqUrGZkOMeYOPwuf5kpenwpgUq9RaOq3fQlhxAjMu3Q",
    "Titans": "1S3J0PwsASL1dH5ggkp9aER264nBNF134VIw3ZAqH8Ug",
}
EXPORT_URL = "https://docs.google.com/spreadsheets/d/{}/export?format=xlsx"
STATE_JSON = state_file("rr_s8_schedule.json")
RESULTS_STATE = state_file("rr_results_state.json")
DEFAULT_RESULTS_CHANNEL = 1527787044514824373  # #match-results in the Pentathletes server
DIV_COLORS = {"Challengers": 0x3498DB, "Legends": 0x9B59B6, "Titans": 0xE74C3C}
DIV_EMOJI = {"Challengers": "🔵", "Legends": "🟣", "Titans": "🔴"}

_HERE = os.path.dirname(os.path.abspath(__file__))
with open(os.path.join(_HERE, "rr_teams.json"), encoding="utf-8") as f:
    TEAM_MAP = json.load(f)  # division -> {hash_to_team, names}

MONTH = re.compile(r"(January|February|March|April|May|June|July|August|September|October|November|December)", re.I)
TIME = re.compile(r"\d\s*(pm|am)", re.I)
SECTION = re.compile(r"(?i)^(saturday|sunday)( [ab])?$")
SCHED_TABS = ["Pre-Season", "Split 1 Schedule", "Major 1", "Split 2 Schedule", "Major 2",
              "League Championship Group Stage", "League Championship Swiss Stage",
              "League Championship Main Event"]
REPORT_TABS = {"Pre-Season": "Pre-Season Game Reports", "Split 1": "Split 1 Game Reports",
               "Major 1": "Major 1 Game Reports", "Split 2": "Split 2 Game Reports",
               "Major 2": "Major 2 Game Reports",
               "League Championship — Group Stage": "Group Stage Game Reports",
               "League Championship — Swiss Stage": "Swiss Stage Game Reports",
               "League Championship — Main Event": "Main Event Game Reports"}
# Standings Automation block labels -> canonical stage names (S7/Auto blocks skipped)
STANDINGS_LABEL = {"Preseason": "Pre-Season", "Split 1": "Split 1", "Major 1": "Major 1",
                   "Split 2": "Split 2", "Major 2": "Major 2",
                   "Group Stage": "League Championship — Group Stage",
                   "Swiss Stage": "League Championship — Swiss Stage",
                   "Main Event": "League Championship — Main Event"}
STAGE_LABEL = {"Pre-Season": "Pre-Season", "Split 1 Schedule": "Split 1", "Major 1": "Major 1",
               "Split 2 Schedule": "Split 2", "Major 2": "Major 2",
               "League Championship Group Stage": "League Championship — Group Stage",
               "League Championship Swiss Stage": "League Championship — Swiss Stage",
               "League Championship Main Event": "League Championship — Main Event"}


# ── xlsx parsing (pure functions, no Discord) ────────────────────────────────

def _drawing_anchors(z: zipfile.ZipFile) -> dict:
    """sheet name -> [{row, col, img}] from the workbook's drawing XML."""
    names = set(z.namelist())
    wbxml = z.read("xl/workbook.xml").decode()
    wbrels = z.read("xl/_rels/workbook.xml.rels").decode()
    rid2target = dict(re.findall(r'Id="([^"]+)"[^>]*Target="([^"]+)"', wbrels))
    sheets = re.findall(r'<sheet[^>]*name="([^"]+)"[^>]*r:id="(rId\d+)"', wbxml)
    anchors = {}
    for sname, rid in sheets:
        sfile = rid2target[rid].split("/")[-1]
        relp = f"xl/worksheets/_rels/{sfile}.rels"
        if relp not in names:
            continue
        m = re.search(r'Target="\.\./(drawings/drawing\d+\.xml)"', z.read(relp).decode())
        if not m:
            continue
        dxml = z.read("xl/" + m.group(1)).decode()
        drelp = "xl/drawings/_rels/" + m.group(1).split("/")[-1] + ".rels"
        if drelp not in names:
            continue
        drid2img = dict(re.findall(r'Id="([^"]+)"[^>]*Target="\.\./(media/[^"]+)"', z.read(drelp).decode()))
        items = []
        for am in re.finditer(r"<xdr:oneCellAnchor>.*?</xdr:oneCellAnchor>", dxml, re.S):
            blk = am.group(0)
            col = re.search(r"<xdr:col>(\d+)</xdr:col>", blk)
            row = re.search(r"<xdr:row>(\d+)</xdr:row>", blk)
            rid_m = re.search(r'r:embed="(rId\d+)"', blk)
            if col and row and rid_m and rid_m.group(1) in drid2img:
                items.append({"row": int(row.group(1)) + 1, "col": int(col.group(1)) + 1,
                              "img": drid2img[rid_m.group(1)].split("/")[-1]})
        if items:
            anchors[sname] = items
    return anchors


def _media_hashes(z: zipfile.ZipFile) -> dict:
    return {n.split("/")[-1]: hashlib.md5(z.read(n)).hexdigest()[:10]
            for n in z.namelist() if n.startswith("xl/media/")}


def _decode_result(m1, m2, t1: str, t2: str) -> dict | None:
    """Two mark-cell values ('W'/'FF'/games-won number) -> result dict or None."""
    def norm(v):
        if isinstance(v, (int, float)):
            return float(v)
        if isinstance(v, str):
            s = v.strip().upper()
            if s in ("W", "FF"):
                return s
            try:
                return float(s)
            except ValueError:
                return None
        return None
    a, b = norm(m1), norm(m2)
    if a is None and b is None:
        return None
    if isinstance(a, float) and isinstance(b, float):
        if a == b:
            return None
        return {"score": [int(a), int(b)], "winner": t1 if a > b else t2, "ff": False}
    if a == "W" or b == "FF":
        return {"score": None, "winner": t1, "ff": True}
    if b == "W" or a == "FF":
        return {"score": None, "winner": t2, "ff": True}
    return None


def _parse_reports(wb, anchors: dict, img_hash: dict, division: str) -> dict:
    """stage -> {(date, frozenset(team pair)): result} from the Game Reports tabs.

    Reports mirror the schedule grid: vertical banner pairs per lane; each
    team row has a mark cell 4 columns left of the banner ('W', 'FF', or the
    number of series games won).
    """
    hash2team = TEAM_MAP[division]["hash_to_team"]
    ignore = set(TEAM_MAP.get("ignore_hashes", []))
    out = {}
    for stage, tab in REPORT_TABS.items():
        items = anchors.get(tab, [])
        if not items or tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        cells, dates = {}, []
        for row in ws.iter_rows(max_row=250):
            for c in row:
                if c.value is None or str(c.value).strip() == "":
                    continue
                cells[(c.row, c.column)] = c.value
                if isinstance(c.value, str) and MONTH.search(c.value):
                    dates.append((c.row, c.column, c.value.strip()))
        hdr_rows = sorted({d[0] for d in dates})

        def sec_start(r):
            prior = [h for h in hdr_rows if h <= r]
            return prior[-1] if prior else 0

        lanes = defaultdict(list)
        for it in sorted(items, key=lambda x: (x["row"], x["col"])):
            h = img_hash.get(it["img"], "")
            if h in ignore:
                continue
            lanes[(sec_start(it["row"]), it["col"])].append((it["row"], hash2team.get(h)))
        results = {}
        for (sec, col), slots in lanes.items():
            sec_dates = [d for d in dates if d[0] == sec] or dates
            slots = sorted(s for s in slots if s[1] is not None)
            for i in range(0, len(slots) - 1, 2):
                (r1, t1), (r2, t2) = slots[i], slots[i + 1]
                date = min(sec_dates, key=lambda d: abs(d[1] - col))[2] if sec_dates else "?"
                res = _decode_result(cells.get((r1, col - 4)), cells.get((r2, col - 4)), t1, t2)
                if res:
                    res["pair"] = [t1, t2]  # score is in this order
                    results[(date, frozenset((t1, t2)))] = res
        if results:
            out[stage] = results
    return out


def _parse_standings(wb, anchors: dict, img_hash: dict, division: str) -> dict:
    """canonical stage -> [{team, pts, w, l, gw, gl, gf, ga}] from Standings Automation."""
    hash2team = TEAM_MAP[division]["hash_to_team"]
    if "Standings Automation" not in wb.sheetnames:
        return {}
    ws = wb["Standings Automation"]
    cells, headers = {}, []
    for row in ws.iter_rows(max_row=300):
        for c in row:
            if c.value is None or str(c.value).strip() == "":
                continue
            cells[(c.row, c.column)] = c.value
            if c.column == 1 and isinstance(c.value, str):
                headers.append((c.row, c.value.strip()))
    team_rows = {it["row"]: it["img"] for it in anchors.get("Standings Automation", [])
                 if it["col"] == 2}
    out = {}
    for i, (hrow, label) in enumerate(headers):
        stage = STANDINGS_LABEL.get(label)
        if stage is None:  # skips "* Auto" duplicates and prior-season blocks
            continue
        end = headers[i + 1][0] if i + 1 < len(headers) else hrow + 30
        rows = []
        for r in range(hrow + 1, end):
            img = team_rows.get(r)
            if not img:
                continue
            team = hash2team.get(img_hash.get(img, ""))
            if not team:
                continue

            def num(col):
                v = cells.get((r, col), 0)
                return float(v) if isinstance(v, (int, float)) else 0.0
            rows.append({"team": team, "pts": num(3), "w": num(4), "l": num(5),
                         "gw": num(7), "gl": num(8), "gf": num(10), "ga": num(11)})
        if rows and any(row["w"] or row["l"] for row in rows):
            out[stage] = rows
    return out


def parse_workbook(xlsx_bytes: bytes, division: str) -> tuple[dict, list]:
    """Returns ({stages for this division}, [unknown banner hashes])."""
    import openpyxl  # heavy import; only needed during sync
    hash2team = TEAM_MAP[division]["hash_to_team"]
    ignore = set(TEAM_MAP.get("ignore_hashes", []))
    z = zipfile.ZipFile(io.BytesIO(xlsx_bytes))
    anchors = _drawing_anchors(z)
    img_hash = _media_hashes(z)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    unknown = set()
    stages = []
    for tab in SCHED_TABS:
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        dates, times, sections = [], [], []
        for row in ws.iter_rows(max_row=200):
            for c in row:
                v = c.value
                if not isinstance(v, str) or not v.strip():
                    continue
                v = v.strip()
                if MONTH.search(v):
                    dates.append((c.row, c.column, v))
                elif TIME.search(v) or v == "TBD":
                    times.append((c.row, c.column, v))
                elif SECTION.fullmatch(v):
                    sections.append((c.row, c.column, v))
        matches = []
        items = anchors.get(tab, [])
        if items:
            hdr_rows = sorted({d[0] for d in dates})

            def sec_start(r):
                prior = [h for h in hdr_rows if h <= r]
                return prior[-1] if prior else 0

            lanes = defaultdict(list)
            for it in sorted(items, key=lambda x: (x["row"], x["col"])):
                h = img_hash.get(it["img"], "")
                if h in ignore:
                    continue
                team = hash2team.get(h)
                if team is None:
                    unknown.add(h)
                lanes[(sec_start(it["row"]), it["col"])].append((it["row"], team))
            for (sec, col), slots in sorted(lanes.items()):
                sec_dates = [d for d in dates if d[0] == sec] or dates
                sec_times = [t for t in times if sec <= t[0] < sec + 60]
                slots = sorted(s for s in slots if s[1] is not None)
                for i in range(0, len(slots) - 1, 2):
                    (r1, t1), (r2, t2) = slots[i], slots[i + 1]
                    date = min(sec_dates, key=lambda d: abs(d[1] - col))[2] if sec_dates else "?"
                    tl = [t for t in sec_times if t[0] <= r2 + 1]
                    time_lbl = min(tl, key=lambda t: abs(t[0] - r1))[2] if tl else "TBD"
                    sl = [s for s in sections if s[0] <= r1 + 2]
                    day = max(sl, key=lambda s: s[0])[2].title() if sl else None
                    matches.append({"date": date, "time": time_lbl, "day": day,
                                    "teams": [t1, t2], "_row": r1, "_col": col})
        matches.sort(key=lambda m: (m["_col"], m["_row"]))
        for m in matches:
            m.pop("_row"), m.pop("_col")
        stages.append({"stage": STAGE_LABEL[tab],
                       "dates": list(dict.fromkeys(d[2] for d in dates)),
                       "matches": matches, "scheduled": bool(matches)})
    reports = _parse_reports(wb, anchors, img_hash, division)
    standings = _parse_standings(wb, anchors, img_hash, division)

    def attach(m, res):
        score = res["score"]
        if score and res["pair"] != m["teams"]:
            score = score[::-1]
        m["result"] = {"score": score, "winner": res["winner"], "ff": res["ff"]}

    def date_key(label):
        m = re.match(r"(\w+) (\d+)", label or "")
        months = ["January", "February", "March", "April", "May", "June", "July",
                  "August", "September", "October", "November", "December"]
        return (months.index(m.group(1)), int(m.group(2))) if m and m.group(1) in months else (99, 99)

    for st in stages:
        stage_results = dict(reports.get(st["stage"], {}))
        # pass 1: exact scheduled date + team pair
        for m in st["matches"]:
            res = stage_results.pop((m["date"], frozenset(m["teams"])), None)
            if res:
                attach(m, res)
        # pass 2: makeup games are reported under the date actually played, so
        # match leftover results to unresolved meetings of the same pair in
        # chronological order
        leftovers = defaultdict(list)
        for (rdate, pair), res in sorted(stage_results.items(), key=lambda kv: date_key(kv[0][0])):
            leftovers[pair].append(res)
        for m in sorted((m for m in st["matches"] if "result" not in m),
                        key=lambda m: date_key(m["date"])):
            q = leftovers.get(frozenset(m["teams"]))
            if q:
                attach(m, q.pop(0))
    wb.close()
    return {"teams": TEAM_MAP[division]["names"], "stages": stages,
            "standings": standings}, sorted(unknown)


def _match_counts(data: dict) -> dict:
    return {div: sum(len(st["matches"]) for st in d["stages"])
            for div, d in data["divisions"].items()}


def _all_results(data: dict):
    """Yields (division, stage, match) for every completed match."""
    for div, d in data["divisions"].items():
        for st in d["stages"]:
            for m in st["matches"]:
                if m.get("result"):
                    yield div, st["stage"], m


def _result_key(div: str, stage: str, m: dict) -> str:
    return "|".join([div, stage, m["date"]] + sorted(m["teams"]))


# ── cog ──────────────────────────────────────────────────────────────────────

class RRSync(commands.Cog):
    def __init__(self, bot: commands.Bot):
        self.bot = bot
        self.last_result = "never run"
        self.state = self._load_results_state()
        self.sync_loop.start()

    def cog_unload(self):
        self.sync_loop.cancel()

    # ── result announcements ─────────────────────────────────────────────────

    def _load_results_state(self) -> dict:
        try:
            with open(RESULTS_STATE, encoding="utf-8") as f:
                return json.load(f)
        except (OSError, ValueError):
            # First run: mark every already-known result as announced so we
            # don't flood the channel with old scores.
            state = {"channel_id": DEFAULT_RESULTS_CHANNEL,
                     "announced": [_result_key(d, s, m) for d, s, m in _all_results(rr_schedule.DATA)]}
            self._save_results_state(state)
            return state

    def _save_results_state(self, state: dict | None = None):
        tmp = RESULTS_STATE + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(state or self.state, f)
        os.replace(tmp, RESULTS_STATE)

    def _new_results(self) -> dict:
        """division -> [(stage, match), ...] not announced yet."""
        announced = set(self.state["announced"])
        fresh = defaultdict(list)
        for div, stage, m in _all_results(rr_schedule.DATA):
            if _result_key(div, stage, m) not in announced:
                fresh[div].append((stage, m))
        return fresh

    async def _announce_results(self) -> int:
        channel_id = self.state.get("channel_id")
        if not channel_id:
            return 0
        fresh = self._new_results()
        if not fresh:
            return 0
        channel = self.bot.get_channel(channel_id)
        if channel is None:
            try:
                channel = await self.bot.fetch_channel(channel_id)
            except discord.HTTPException:
                print(f"[rr_sync] results channel {channel_id} not reachable")
                return 0
        count = 0
        for div, entries in fresh.items():
            embed = discord.Embed(title=f"{DIV_EMOJI.get(div, '')} RRS8 {div.upper()} — MATCH RESULTS",
                                  color=DIV_COLORS.get(div, 0x5865F2))
            by_stage = defaultdict(list)
            for stage, m in entries:
                by_stage[stage].append(m)
            for stage, ms in by_stage.items():
                lines = "\n".join(f"**{rr_schedule.format_result(m)}**  ·  {m['date']}"
                                  + (" (forfeit)" if m["result"]["ff"] else "") for m in ms)
                embed.add_field(name=stage, value=lines[:1024], inline=False)
            embed.set_footer(text="Use /standings and /powerrankings for the bigger picture")
            await channel.send(embed=embed)
            for stage, m in entries:
                self.state["announced"].append(_result_key(div, stage, m))
            count += len(entries)
        self._save_results_state()
        return count

    async def sync_once(self) -> str:
        new_data = {"season": 8, "year": rr_schedule.SEASON_YEAR,
                    "source_updated": datetime.now().strftime("%Y-%m-%d"),
                    "divisions": {}}
        unknown_all = {}
        async with aiohttp.ClientSession() as session:
            for div, sheet_id in SHEET_IDS.items():
                raw = None
                for attempt in (1, 2, 3):
                    try:
                        async with session.get(EXPORT_URL.format(sheet_id),
                                               timeout=aiohttp.ClientTimeout(total=180)) as resp:
                            resp.raise_for_status()
                            raw = await resp.read()
                        break
                    except aiohttp.ClientError:
                        if attempt == 3:
                            raise
                div_data, unknown = await self.bot.loop.run_in_executor(None, parse_workbook, raw, div)
                new_data["divisions"][div] = div_data
                if unknown:
                    unknown_all[div] = unknown

        old_counts = _match_counts(rr_schedule.DATA)
        new_counts = _match_counts(new_data)
        # Guard against a bad download/parse wiping a division that had data
        for div, n in new_counts.items():
            if n == 0 and old_counts.get(div, 0) > 0:
                self.last_result = f"rejected: {div} parsed to 0 matches (had {old_counts[div]})"
                print(f"[rr_sync] {self.last_result}")
                return self.last_result

        old_matches = {d: (v["stages"], v.get("standings")) for d, v in rr_schedule.DATA["divisions"].items()}
        new_matches = {d: (v["stages"], v.get("standings")) for d, v in new_data["divisions"].items()}
        changed = old_matches != new_matches
        if changed:
            tmp = STATE_JSON + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(new_data, f, indent=1)
            os.replace(tmp, STATE_JSON)
            rr_schedule.reload()
            summary = ", ".join(f"{d}: {old_counts.get(d, 0)}→{n}" for d, n in new_counts.items())
            self.last_result = f"updated ({summary})"
        else:
            self.last_result = f"no changes ({', '.join(f'{d}: {n}' for d, n in new_counts.items())})"
        announced = await self._announce_results()
        if announced:
            self.last_result += f" | announced {announced} new result(s)"
        if unknown_all:
            notes = "; ".join(f"{d}: {len(u)} unknown banner(s) {u}" for d, u in unknown_all.items())
            self.last_result += f" | NEW TEAMS NEED MAPPING — {notes}"
        print(f"[rr_sync] {self.last_result}")
        return self.last_result

    @tasks.loop(hours=6)
    async def sync_loop(self):
        try:
            await self.sync_once()
        except Exception as e:
            self.last_result = f"error: {e}"
            print(f"[rr_sync] sync failed: {e}")

    @sync_loop.before_loop
    async def before_sync(self):
        await self.bot.wait_until_ready()

    @app_commands.command(name="setresults",
                          description="Admin: set (or move) the channel where new RRS8 match results get posted")
    @app_commands.describe(channel="Channel for result announcements")
    @app_commands.default_permissions(manage_guild=True)
    async def setresults(self, interaction: discord.Interaction, channel: discord.TextChannel):
        self.state["channel_id"] = channel.id
        self._save_results_state()
        await interaction.response.send_message(
            f"New RRS8 match results will be posted in {channel.mention}.", ephemeral=True)

    @app_commands.command(name="syncschedule",
                          description="Admin: re-pull the RRS8 schedules from the official sheets now")
    @app_commands.default_permissions(manage_guild=True)
    async def syncschedule(self, interaction: discord.Interaction):
        await interaction.response.defer(ephemeral=True)
        try:
            result = await self.sync_once()
        except Exception as e:
            result = f"error: {e}"
        await interaction.followup.send(f"Schedule sync: {result}")


async def setup(bot: commands.Bot):
    await bot.add_cog(RRSync(bot))
